#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
利润计算模块 - 计算利润并生成Excel
"""

import pandas as pd
from matcher import find_matching_price


def calculate_profit_and_generate_excel(orders, price_table, purchase_table, output_file):
    """
    计算利润并生成Excel文件
    
    参数:
        orders (list): 订单列表
        price_table (dict): 定价表
        purchase_table (dict): 进价表
        output_file (str): 输出文件路径
    
    返回:
        float: 总利润
    """
    # 数据验证
    if not orders:
        raise ValueError("订单列表为空，请检查文字列表格式")
    
    # 构建数据列表
    data = []
    
    for order in orders:
        # 验证订单数据完整性
        if '单位' not in order:
            order['单位'] = '未指定单位'
        if '菜品' not in order:
            continue  # 跳过无效订单
        if '数量' not in order:
            order['数量'] = 0
        unit = order['单位']
        vegetable = order['菜品']
        quantity = order['数量']

        # 使用更强的模糊匹配逻辑
        from matcher import fuzzy_find_price, find_matching_cost_price
        selling_price = fuzzy_find_price(vegetable, price_table)
        # 进价使用专门的进价匹配函数，支持规格通用匹配
        purchase_price = find_matching_cost_price(vegetable, purchase_table)

        # 计算利润
        if selling_price != "000000" and purchase_price != "000000":
            unit_profit = selling_price - purchase_price
            total_profit = unit_profit * quantity
        else:
            unit_profit = "000000"
            total_profit = "000000"

        # 计算总金额（定价*数量），如果定价为000000则为000000
        if selling_price != "000000":
            total_amount = selling_price * quantity
        else:
            total_amount = "000000"

        data.append({
            '单位': unit,
            '菜品': vegetable,
            '数量': quantity,
            '定价': selling_price,
            '进价': purchase_price,
            '单品利润': unit_profit,
            '总利润': total_profit,
            '总金额': total_amount
        })
    
    # 创建DataFrame
    df = pd.DataFrame(data)
    
    # 按单位分组排序
    df = df.sort_values(by=['单位', '菜品'])
    
    # 计算总利润
    total_profit_sum = 0
    for profit in df['总利润']:
        if profit != "000000":
            total_profit_sum += profit
    
    # 计算总金额合计
    total_amount_sum = 0
    if '总金额' in df.columns:
        for amount in df['总金额']:
            if amount != "000000":
                total_amount_sum += amount

    # 添加总计行
    total_row = pd.DataFrame([{
        '单位': '总计',
        '菜品': '',
        '数量': '',
        '定价': '',
        '进价': '',
        '单品利润': '',
        '总利润': total_profit_sum,
        '总金额': total_amount_sum
    }])
    df = pd.concat([df, total_row], ignore_index=True)
    
    # 保存到Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='利润表')
        
        # 获取工作表
        workbook = writer.book
        worksheet = writer.sheets['利润表']
        
        # 设置列宽
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 10
        worksheet.column_dimensions['D'].width = 10
        worksheet.column_dimensions['E'].width = 10
        worksheet.column_dimensions['F'].width = 12
        worksheet.column_dimensions['G'].width = 12
        worksheet.column_dimensions['H'].width = 14

    # 设置表头样式
    from openpyxl.styles import Font, Alignment, PatternFill

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')

    for cell in worksheet[1]:
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 设置总计行样式
    last_row = len(df) + 1
    for cell in worksheet[last_row]:
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    
    return total_profit_sum


if __name__ == "__main__":
    # 测试
    orders = [
        {'单位': '一中五食堂', '菜品': '胡萝卜', '数量': 20},
        {'单位': '一中五食堂', '菜品': '尖椒', '数量': 20},
    ]
    
    price_table = {'胡萝卜': 5, '尖椒': 8}
    purchase_table = {'胡萝卜': 3, '尖椒': 5}
    
    total = calculate_profit_and_generate_excel(orders, price_table, purchase_table, "测试.xlsx")
    print(f"总利润: {total}")
